"""
Rutas de Administración
=======================
Dashboard admin, gestión de usuarios y configuración.
"""
from flask import Blueprint, render_template, session, jsonify, request, make_response, send_file
from werkzeug.security import generate_password_hash
from datetime import datetime
import sqlite3
import io

from models.database import get_db
from utils.helpers import admin_required, login_required

# Crear el Blueprint
admin_bp = Blueprint('admin', __name__, url_prefix='/admin')


# ============================================
# DASHBOARD ADMIN
# ============================================

@admin_bp.route("")
@admin_required
def admin_dashboard():
    """Dashboard principal del administrador"""
    try:
        db = get_db()
        cursor = db.cursor()
        
        hoy = datetime.now().strftime("%Y-%m-%d")
        mes_actual = datetime.now().strftime("%Y-%m")
        
        # Ingresos del día
        cursor.execute("""
            SELECT 
                IFNULL(SUM(CASE WHEN metodo_pago = 'efectivo' THEN monto ELSE 0 END), 0) as efectivo,
                IFNULL(SUM(CASE WHEN metodo_pago = 'yape' THEN monto ELSE 0 END), 0) as yape,
                IFNULL(SUM(monto), 0) as total
            FROM movimientos_caja
            WHERE date(fecha_movimiento) = ?
        """, (hoy,))
        ingresos_dia = cursor.fetchone()
        
        # Ingresos del mes
        cursor.execute("""
            SELECT 
                IFNULL(SUM(CASE WHEN metodo_pago = 'efectivo' THEN monto ELSE 0 END), 0) as efectivo,
                IFNULL(SUM(CASE WHEN metodo_pago = 'yape' THEN monto ELSE 0 END), 0) as yape,
                IFNULL(SUM(monto), 0) as total
            FROM movimientos_caja
            WHERE strftime('%Y-%m', fecha_movimiento) = ?
        """, (mes_actual,))
        ingresos_mes = cursor.fetchone()
        
        # Total histórico
        cursor.execute("""
            SELECT 
                IFNULL(SUM(CASE WHEN metodo_pago = 'efectivo' THEN monto ELSE 0 END), 0) as efectivo,
                IFNULL(SUM(CASE WHEN metodo_pago = 'yape' THEN monto ELSE 0 END), 0) as yape,
                IFNULL(SUM(monto), 0) as total
            FROM movimientos_caja
        """)
        total_historico = cursor.fetchone()
        
        # Estadísticas
        cursor.execute("SELECT COUNT(*) FROM entradas WHERE salio = 0")
        autos_en_cochera = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM clientes")
        total_clientes = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM trabajadores WHERE activo = 1")
        total_trabajadores = cursor.fetchone()[0]
        
        # Ingresos de la semana
        cursor.execute("""
            SELECT 
                date(fecha_movimiento) as fecha,
                SUM(monto) as total
            FROM movimientos_caja
            WHERE fecha_movimiento >= date('now', '-7 days')
            GROUP BY date(fecha_movimiento)
            ORDER BY fecha
        """)
        ingresos_semana = cursor.fetchall()

        return render_template(
            "admin_dashboard.html",
            nombre=session["nombre"],
            ingresos_dia=dict(ingresos_dia),
            ingresos_mes=dict(ingresos_mes),
            total_historico=dict(total_historico),
            autos_en_cochera=autos_en_cochera,
            total_clientes=total_clientes,
            total_trabajadores=total_trabajadores,
            ingresos_semana=[dict(i) for i in ingresos_semana]
        )

    except Exception as e:
        print(f"Error en admin dashboard: {e}")
        return "Error al cargar dashboard", 500


# ============================================
# TURNO ACTIVO
# ============================================

@admin_bp.route("/turno_activo")
@admin_required
def turno_activo():
    """Obtiene el turno activo actual"""
    try:
        db = get_db()
        cursor = db.cursor()

        cursor.execute("""
            SELECT t.id, t.fecha_inicio, tr.nombre as trabajador
            FROM turnos t
            JOIN trabajadores tr ON t.trabajador_id = tr.id
            WHERE t.estado = 'abierto'
            ORDER BY t.fecha_inicio DESC
            LIMIT 1
        """)

        turno = cursor.fetchone()

        if turno:
            return jsonify({
                "ok": True,
                "turno": {
                    "id": turno["id"],
                    "trabajador": turno["trabajador"],
                    "fecha_inicio": turno["fecha_inicio"]
                }
            })
        else:
            return jsonify({"ok": True, "turno": None})

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


# ============================================
# GESTIÓN DE USUARIOS
# ============================================

@admin_bp.route("/usuarios")
@admin_required
def listar_usuarios():
    """Lista todos los usuarios"""
    db = get_db()
    cursor = db.cursor()
    
    cursor.execute("""
        SELECT id, nombre, usuario, rol, activo, fecha_creacion
        FROM trabajadores
        ORDER BY rol DESC, nombre
    """)
    
    usuarios = [dict(u) for u in cursor.fetchall()]
    
    return jsonify({"ok": True, "usuarios": usuarios})


@admin_bp.route("/usuarios/crear", methods=["POST"])
@admin_required
def crear_usuario():
    """Crea un nuevo usuario"""
    data = request.json
    
    if not data.get("nombre") or not data.get("usuario") or not data.get("password"):
        return jsonify({"ok": False, "error": "Todos los campos son requeridos"})
    
    db = get_db()
    cursor = db.cursor()
    
    try:
        password_hash = generate_password_hash(data["password"])
        cursor.execute("""
            INSERT INTO trabajadores (nombre, usuario, password, rol)
            VALUES (?, ?, ?, ?)
        """, (
            data["nombre"],
            data["usuario"].lower().strip(),
            password_hash,
            data.get("rol", "trabajador")
        ))
        
        db.commit()
        return jsonify({"ok": True, "mensaje": "Usuario creado exitosamente"})
        
    except sqlite3.IntegrityError:
        return jsonify({"ok": False, "error": "El usuario ya existe"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@admin_bp.route("/usuarios/editar", methods=["POST"])
@admin_required
def editar_usuario():
    """Edita un usuario existente"""
    data = request.json
    
    if not data.get("id"):
        return jsonify({"ok": False, "error": "ID requerido"})
    
    db = get_db()
    cursor = db.cursor()
    
    try:
        if data.get("password"):
            password_hash = generate_password_hash(data["password"])
            cursor.execute("""
                UPDATE trabajadores
                SET nombre = ?, usuario = ?, password = ?, rol = ?, activo = ?
                WHERE id = ?
            """, (
                data["nombre"],
                data["usuario"].lower().strip(),
                password_hash,
                data.get("rol", "trabajador"),
                1 if data.get("activo", True) else 0,
                data["id"]
            ))
        else:
            cursor.execute("""
                UPDATE trabajadores
                SET nombre = ?, usuario = ?, rol = ?, activo = ?
                WHERE id = ?
            """, (
                data["nombre"],
                data["usuario"].lower().strip(),
                data.get("rol", "trabajador"),
                1 if data.get("activo", True) else 0,
                data["id"]
            ))
        
        db.commit()
        return jsonify({"ok": True, "mensaje": "Usuario actualizado"})
        
    except sqlite3.IntegrityError:
        return jsonify({"ok": False, "error": "El usuario ya existe"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@admin_bp.route("/usuarios/eliminar/<int:id>", methods=["DELETE"])
@admin_required
def eliminar_usuario(id):
    """Desactiva un usuario (no elimina)"""
    if id == session["trabajador_id"]:
        return jsonify({"ok": False, "error": "No puedes eliminarte a ti mismo"})
    
    db = get_db()
    cursor = db.cursor()
    
    try:
        cursor.execute("UPDATE trabajadores SET activo = 0 WHERE id = ?", (id,))
        db.commit()
        return jsonify({"ok": True, "mensaje": "Usuario desactivado"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


# ============================================
# CONFIGURACIÓN
# ============================================

@admin_bp.route("/configuracion")
@admin_required
def obtener_configuracion():
    """Obtiene toda la configuración"""
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT * FROM configuracion")
    config = [dict(c) for c in cursor.fetchall()]
    return jsonify({"ok": True, "configuracion": config})


@admin_bp.route("/configuracion", methods=["POST"])
@admin_required
def guardar_configuracion():
    """Guarda la configuración"""
    data = request.json
    
    db = get_db()
    cursor = db.cursor()
    
    try:
        for clave, valor in data.items():
            cursor.execute("""
                UPDATE configuracion SET valor = ? WHERE clave = ?
            """, (valor, clave))
        
        db.commit()
        return jsonify({"ok": True, "mensaje": "Configuración guardada"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


# ============================================
# HISTORIAL Y REPORTES
# ============================================

@admin_bp.route("/historial_vehiculos")
@login_required
def historial_vehiculos():
    """Historial completo de vehículos"""
    try:
        db = get_db()
        cursor = db.cursor()
        
        filtro_placa = request.args.get('placa', '').upper()
        filtro_fecha_desde = request.args.get('fecha_desde', '')
        filtro_fecha_hasta = request.args.get('fecha_hasta', '')
        filtro_estado = request.args.get('estado', '')
        pagina = int(request.args.get('pagina', 1))
        por_pagina = int(request.args.get('por_pagina', 50))
        
        query = """
            SELECT 
                e.id,
                c.placa,
                c.nombre as cliente,
                c.celular,
                e.fecha_entrada,
                e.hora_entrada,
                e.fecha_salida,
                e.hora_salida_real,
                e.dias,
                e.precio_dia,
                e.monto,
                e.adelanto,
                e.penalidad,
                e.descuento,
                e.metodo_pago,
                e.pagado,
                e.pago_completo_adelantado,
                e.salio,
                e.observaciones,
                t1.nombre as trabajador_entrada,
                t2.nombre as trabajador_salida,
                CASE
                    WHEN e.salio = 0 THEN MAX(1, CAST((julianday(datetime('now', 'localtime')) - julianday(e.fecha_entrada)) + 0.5 AS INTEGER))
                    ELSE e.dias
                END as dias_reales
            FROM entradas e
            JOIN clientes c ON e.cliente_id = c.id
            LEFT JOIN trabajadores t1 ON e.trabajador_id = t1.id
            LEFT JOIN trabajadores t2 ON e.trabajador_salida_id = t2.id
            WHERE 1=1
        """
        params = []
        
        if filtro_placa:
            query += " AND c.placa LIKE ?"
            params.append(f"%{filtro_placa}%")
        
        if filtro_fecha_desde:
            query += " AND e.fecha_entrada >= ?"
            params.append(filtro_fecha_desde)
        
        if filtro_fecha_hasta:
            query += " AND e.fecha_entrada <= ?"
            params.append(filtro_fecha_hasta)
        
        if filtro_estado == 'en_cochera':
            query += " AND e.salio = 0"
        elif filtro_estado == 'salieron':
            query += " AND e.salio = 1"
        
        # Contar total
        count_query = f"SELECT COUNT(*) FROM ({query})"
        cursor.execute(count_query, params)
        total = cursor.fetchone()[0]
        
        # Ordenar y paginar
        query += " ORDER BY e.fecha_entrada DESC, e.hora_entrada DESC"
        query += f" LIMIT {por_pagina} OFFSET {(pagina - 1) * por_pagina}"
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        
        historial = []
        for r in rows:
            historial.append({
                "id": r["id"],
                "placa": r["placa"],
                "cliente": r["cliente"],
                "celular": r["celular"],
                "fecha_entrada": r["fecha_entrada"],
                "hora_entrada": r["hora_entrada"],
                "fecha_salida": r["fecha_salida"],
                "hora_salida": r["hora_salida_real"],
                "dias": r["dias"],
                "dias_reales": r["dias_reales"],
                "precio_dia": r["precio_dia"],
                "monto": r["monto"],
                "adelanto": r["adelanto"],
                "penalidad": r["penalidad"],
                "descuento": r["descuento"],
                "metodo_pago": r["metodo_pago"],
                "pagado": r["pagado"],
                "pago_completo": r["pago_completo_adelantado"],
                "salio": r["salio"],
                "observaciones": r["observaciones"],
                "trabajador_entrada": r["trabajador_entrada"],
                "trabajador_salida": r["trabajador_salida"]
            })
        
        return jsonify({
            "ok": True,
            "historial": historial,
            "total": total,
            "pagina": pagina,
            "por_pagina": por_pagina,
            "total_paginas": (total + por_pagina - 1) // por_pagina
        })
        
    except Exception as e:
        print(f"Error en historial: {e}")
        return jsonify({"ok": False, "error": str(e)})


@admin_bp.route("/exportar_historial")
@login_required
def exportar_historial():
    """Exporta el historial a Excel"""
    try:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return jsonify({"ok": False, "error": "Módulo openpyxl no instalado"})
        
        db = get_db()
        cursor = db.cursor()
        
        filtro_placa = request.args.get('placa', '').upper()
        filtro_fecha_desde = request.args.get('fecha_desde', '')
        filtro_fecha_hasta = request.args.get('fecha_hasta', '')
        filtro_estado = request.args.get('estado', '')
        
        query = """
            SELECT 
                c.placa,
                c.nombre as cliente,
                c.celular,
                e.fecha_entrada,
                e.hora_entrada,
                e.fecha_salida,
                e.hora_salida_real as hora_salida,
                e.dias,
                e.precio_dia,
                e.monto,
                e.adelanto,
                e.penalidad,
                e.descuento,
                e.metodo_pago,
                CASE WHEN e.salio = 1 THEN 'Sí' ELSE 'No' END as salio,
                CASE WHEN e.pagado = 1 THEN 'Sí' ELSE 'No' END as pagado,
                t1.nombre as trabajador_entrada,
                t2.nombre as trabajador_salida,
                e.observaciones
            FROM entradas e
            JOIN clientes c ON e.cliente_id = c.id
            LEFT JOIN trabajadores t1 ON e.trabajador_id = t1.id
            LEFT JOIN trabajadores t2 ON e.trabajador_salida_id = t2.id
            WHERE 1=1
        """
        params = []
        
        if filtro_placa:
            query += " AND c.placa LIKE ?"
            params.append(f"%{filtro_placa}%")
        if filtro_fecha_desde:
            query += " AND e.fecha_entrada >= ?"
            params.append(filtro_fecha_desde)
        if filtro_fecha_hasta:
            query += " AND e.fecha_entrada <= ?"
            params.append(filtro_fecha_hasta)
        if filtro_estado == 'en_cochera':
            query += " AND e.salio = 0"
        elif filtro_estado == 'salieron':
            query += " AND e.salio = 1"
        
        query += " ORDER BY e.fecha_entrada DESC"
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        
        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Historial"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = ['Placa', 'Cliente', 'Celular', 'F. Entrada', 'H. Entrada', 
                   'F. Salida', 'H. Salida', 'Días', 'Precio/Día', 'Monto Total',
                   'Adelanto', 'Penalidad', 'Descuento', 'Método Pago', 'Salió', 
                   'Pagado', 'Trabajador Entrada', 'Trabajador Salida', 'Observaciones']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(dict(row).values(), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
        
        # Ajustar anchos
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 30)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=historial_cochera_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        print(f"Error exportando: {e}")
        return jsonify({"ok": False, "error": str(e)})


@admin_bp.route("/reportes_turnos")
@admin_required
def reportes_turnos():
    """Reportes de turnos de todos los trabajadores"""
    try:
        db = get_db()
        cursor = db.cursor()

        filtro_trabajador = request.args.get('trabajador_id', '')
        filtro_fecha_desde = request.args.get('fecha_desde', '')
        filtro_fecha_hasta = request.args.get('fecha_hasta', '')
        pagina = int(request.args.get('pagina', 1))
        por_pagina = int(request.args.get('por_pagina', 20))

        query = """
            SELECT
                t.id,
                t.trabajador_id,
                tr.nombre as trabajador,
                t.fecha_inicio,
                t.fecha_fin,
                t.estado,
                IFNULL(t.total_efectivo, 0) as total_efectivo,
                IFNULL(t.total_yape, 0) as total_yape,
                IFNULL(t.total_efectivo, 0) + IFNULL(t.total_yape, 0) as total,
                t.efectivo_declarado,
                IFNULL(t.efectivo_declarado, 0) - IFNULL(t.total_efectivo, 0) as diferencia,
                t.observaciones,
                (SELECT COUNT(*) FROM movimientos_caja m WHERE m.turno_id = t.id) as num_movimientos
            FROM turnos t
            JOIN trabajadores tr ON t.trabajador_id = tr.id
            WHERE 1=1
        """
        params = []

        if filtro_trabajador:
            query += " AND t.trabajador_id = ?"
            params.append(int(filtro_trabajador))
        if filtro_fecha_desde:
            query += " AND date(t.fecha_inicio) >= ?"
            params.append(filtro_fecha_desde)
        if filtro_fecha_hasta:
            query += " AND date(t.fecha_inicio) <= ?"
            params.append(filtro_fecha_hasta)

        count_query = f"SELECT COUNT(*) FROM ({query})"
        cursor.execute(count_query, params)
        total = cursor.fetchone()[0]

        query += " ORDER BY t.fecha_inicio DESC"
        query += f" LIMIT {por_pagina} OFFSET {(pagina - 1) * por_pagina}"

        cursor.execute(query, params)
        turnos = [dict(r) for r in cursor.fetchall()]

        # Lista de trabajadores para el filtro
        cursor.execute("SELECT id, nombre FROM trabajadores WHERE activo = 1 ORDER BY nombre")
        trabajadores = [dict(t) for t in cursor.fetchall()]

        return jsonify({
            "ok": True,
            "turnos": turnos,
            "trabajadores": trabajadores,
            "total": total,
            "pagina": pagina,
            "por_pagina": por_pagina,
            "total_paginas": (total + por_pagina - 1) // por_pagina
        })

    except Exception as e:
        print(f"Error en reportes_turnos: {e}")
        return jsonify({"ok": False, "error": str(e)})


@admin_bp.route("/detalle_turno/<int:turno_id>")
@admin_required
def detalle_turno(turno_id):
    """Detalle de movimientos de un turno específico"""
    try:
        db = get_db()
        cursor = db.cursor()

        cursor.execute("""
            SELECT t.*, tr.nombre as trabajador
            FROM turnos t
            JOIN trabajadores tr ON t.trabajador_id = tr.id
            WHERE t.id = ?
        """, (turno_id,))
        turno = cursor.fetchone()
        if not turno:
            return jsonify({"ok": False, "error": "Turno no encontrado"})

        cursor.execute("""
            SELECT
                m.*,
                c.placa,
                c.nombre as cliente
            FROM movimientos_caja m
            LEFT JOIN entradas e ON m.entrada_id = e.id
            LEFT JOIN clientes c ON e.cliente_id = c.id
            WHERE m.turno_id = ?
            ORDER BY m.fecha_movimiento DESC
        """, (turno_id,))
        movimientos = [dict(m) for m in cursor.fetchall()]

        cursor.execute("""
            SELECT
                IFNULL(SUM(CASE WHEN tipo LIKE '%ADELANTO%' OR tipo = 'PAGO_COMPLETO' THEN monto ELSE 0 END), 0) as total_adelantos,
                IFNULL(SUM(CASE WHEN tipo = 'COBRO_SALIDA' THEN monto ELSE 0 END), 0) as total_cobros,
                IFNULL(SUM(CASE WHEN tipo = 'PENALIDAD' THEN monto ELSE 0 END), 0) as total_penalidades
            FROM movimientos_caja WHERE turno_id = ?
        """, (turno_id,))
        desglose = dict(cursor.fetchone())

        return jsonify({
            "ok": True,
            "turno": dict(turno),
            "movimientos": movimientos,
            "desglose": desglose
        })

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@admin_bp.route("/detalle_movimiento/<int:id>")
@login_required
def detalle_movimiento(id):
    """Obtiene el detalle de un movimiento de caja"""
    try:
        db = get_db()
        cursor = db.cursor()

        cursor.execute("""
            SELECT 
                m.*,
                e.fecha_entrada,
                e.hora_entrada,
                e.fecha_salida,
                e.hora_salida_real,
                e.dias,
                e.precio_dia,
                e.monto as monto_total,
                e.adelanto,
                e.penalidad,
                e.descuento,
                e.observaciones,
                e.pago_completo_adelantado,
                c.placa,
                c.nombre as cliente,
                c.celular,
                t1.nombre as trabajador_entrada,
                t2.nombre as trabajador_salida,
                t3.nombre as trabajador_movimiento
            FROM movimientos_caja m
            LEFT JOIN entradas e ON m.entrada_id = e.id
            LEFT JOIN clientes c ON e.cliente_id = c.id
            LEFT JOIN trabajadores t1 ON e.trabajador_id = t1.id
            LEFT JOIN trabajadores t2 ON e.trabajador_salida_id = t2.id
            LEFT JOIN trabajadores t3 ON m.trabajador_id = t3.id
            WHERE m.id = ?
        """, (id,))

        mov = cursor.fetchone()

        if not mov:
            return jsonify({"ok": False, "error": "Movimiento no encontrado"})

        return jsonify({
            "ok": True,
            "movimiento": dict(mov)
        })

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


# ============================================
# GESTIÓN DE CLIENTES
# ============================================

@admin_bp.route("/clientes")
@admin_required
def listar_clientes():
    """Lista todos los clientes con estadísticas"""
    try:
        db = get_db()
        cursor = db.cursor()

        busqueda = request.args.get('busqueda', '').strip().upper()
        pagina = int(request.args.get('pagina', 1))
        por_pagina = int(request.args.get('por_pagina', 50))

        query = """
            SELECT
                c.id,
                c.placa,
                c.nombre,
                c.celular,
                c.precio_dia,
                c.fecha_actualizacion,
                COUNT(e.id) as total_visitas,
                MAX(e.fecha_entrada) as ultima_visita,
                SUM(CASE WHEN e.salio = 0 THEN 1 ELSE 0 END) as entradas_activas
            FROM clientes c
            LEFT JOIN entradas e ON c.id = e.cliente_id
            WHERE 1=1
        """
        params = []

        if busqueda:
            query += " AND (c.placa LIKE ? OR c.nombre LIKE ?)"
            params.extend([f"%{busqueda}%", f"%{busqueda}%"])

        query += " GROUP BY c.id"

        # Contar total
        count_query = f"SELECT COUNT(*) FROM ({query})"
        cursor.execute(count_query, params)
        total = cursor.fetchone()[0]

        # Ordenar y paginar
        query += " ORDER BY c.fecha_actualizacion DESC"
        query += f" LIMIT {por_pagina} OFFSET {(pagina - 1) * por_pagina}"

        cursor.execute(query, params)
        clientes = [dict(c) for c in cursor.fetchall()]

        return jsonify({
            "ok": True,
            "clientes": clientes,
            "total": total,
            "pagina": pagina,
            "por_pagina": por_pagina,
            "total_paginas": (total + por_pagina - 1) // por_pagina
        })

    except Exception as e:
        print(f"Error en listar_clientes: {e}")
        return jsonify({"ok": False, "error": str(e)})


@admin_bp.route("/clientes/crear", methods=["POST"])
@admin_required
def crear_cliente():
    """Crear nuevo cliente"""
    data = request.json

    if not data.get("placa"):
        return jsonify({"ok": False, "error": "La placa es requerida"})

    try:
        db = get_db()
        cursor = db.cursor()

        placa = data["placa"].upper().strip()

        # Verificar que no exista
        cursor.execute("SELECT id FROM clientes WHERE placa = ?", (placa,))
        if cursor.fetchone():
            return jsonify({"ok": False, "error": "Ya existe un cliente con esa placa"})

        nombre = data.get("nombre", "").strip() or "Sin nombre"

        cursor.execute("""
            INSERT INTO clientes (placa, nombre, celular, precio_dia, fecha_actualizacion)
            VALUES (?, ?, ?, ?, datetime('now', 'localtime'))
        """, (
            placa,
            nombre,
            data.get("celular", ""),
            float(data.get("precio_dia", 10))
        ))

        db.commit()
        return jsonify({"ok": True, "mensaje": "Cliente creado exitosamente", "id": cursor.lastrowid})

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@admin_bp.route("/clientes/editar", methods=["POST"])
@admin_required
def editar_cliente():
    """Editar cliente existente"""
    data = request.json

    if not data.get("id"):
        return jsonify({"ok": False, "error": "ID requerido"})

    try:
        db = get_db()
        cursor = db.cursor()

        # Verificar que exista
        cursor.execute("SELECT id FROM clientes WHERE id = ?", (data["id"],))
        if not cursor.fetchone():
            return jsonify({"ok": False, "error": "Cliente no encontrado"})

        # Si se cambia la placa, verificar que no exista otra
        if data.get("placa"):
            placa = data["placa"].upper().strip()
            cursor.execute("SELECT id FROM clientes WHERE placa = ? AND id != ?", (placa, data["id"]))
            if cursor.fetchone():
                return jsonify({"ok": False, "error": "Ya existe otro cliente con esa placa"})
        else:
            cursor.execute("SELECT placa FROM clientes WHERE id = ?", (data["id"],))
            placa = cursor.fetchone()["placa"]

        nombre = data.get("nombre", "").strip() or "Sin nombre"

        cursor.execute("""
            UPDATE clientes
            SET placa = ?, nombre = ?, celular = ?, precio_dia = ?, fecha_actualizacion = datetime('now', 'localtime')
            WHERE id = ?
        """, (
            placa,
            nombre,
            data.get("celular", ""),
            float(data.get("precio_dia", 10)),
            data["id"]
        ))

        db.commit()
        return jsonify({"ok": True, "mensaje": "Cliente actualizado"})

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@admin_bp.route("/clientes/eliminar/<int:id>", methods=["DELETE"])
@admin_required
def eliminar_cliente(id):
    """Eliminar cliente (si no tiene entradas activas)"""
    try:
        db = get_db()
        cursor = db.cursor()

        # Verificar que exista
        cursor.execute("SELECT id FROM clientes WHERE id = ?", (id,))
        if not cursor.fetchone():
            return jsonify({"ok": False, "error": "Cliente no encontrado"})

        # Verificar que no tenga entradas activas
        cursor.execute("SELECT COUNT(*) FROM entradas WHERE cliente_id = ? AND salio = 0", (id,))
        if cursor.fetchone()[0] > 0:
            return jsonify({"ok": False, "error": "No se puede eliminar: el cliente tiene vehículos en cochera"})

        # Eliminar entradas históricas primero
        cursor.execute("DELETE FROM entradas WHERE cliente_id = ?", (id,))
        # Eliminar cliente
        cursor.execute("DELETE FROM clientes WHERE id = ?", (id,))

        db.commit()
        return jsonify({"ok": True, "mensaje": "Cliente eliminado"})

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


# ============================================
# BACKUP BASE DE DATOS
# ============================================
@admin_bp.route("/backup_db")
@admin_required
def backup_db():
    """Descarga una copia de la base de datos"""
    import os
    from flask import current_app
    db_path = current_app.config.get('DATABASE_PATH', 'database.db')
    if not os.path.exists(db_path):
        return "Base de datos no encontrada", 404
    fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(
        db_path,
        as_attachment=True,
        download_name=f"cochera_backup_{fecha}.db"
    )
