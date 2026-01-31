from flask import Flask, render_template, request, redirect, session, jsonify, make_response, send_from_directory
import sqlite3
from datetime import datetime, timedelta
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import os
import io

app = Flask(__name__)

# ============================================
# CONFIGURACIÓN PARA PRODUCCIÓN
# ============================================
# En producción, configura esta variable de entorno:
# export SECRET_KEY="tu_clave_super_secreta_y_larga_aqui_123456789"
app.secret_key = os.environ.get('SECRET_KEY', 'dev_key_cambiar_en_produccion_123')

# Configuración de sesión segura
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('PRODUCTION', False)  # True en HTTPS
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=12)


# -------------------------------
# SERVIR MANIFEST.JSON PARA PWA
# -------------------------------
@app.route('/manifest.json')
def manifest():
    return send_from_directory('static', 'manifest.json')


# -------------------------------
# DECORADORES
# -------------------------------
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "trabajador_id" not in session:
            return redirect("/")
        return f(*args, **kwargs)
    return decorated_function


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "trabajador_id" not in session:
            return redirect("/")
        if not session.get("es_admin"):
            return jsonify({"ok": False, "error": "Acceso denegado. Se requiere rol de administrador."}), 403
        return f(*args, **kwargs)
    return decorated_function


# -------------------------------
# CONEXIÓN A LA BASE DE DATOS
# -------------------------------
def conectar_db():
    # En producción puedes usar variable de entorno para la ruta
    db_path = os.environ.get('DATABASE_PATH', 'database.db')
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


# -------------------------------
# CREAR TABLAS
# -------------------------------
def crear_tablas():
    conn = conectar_db()
    cursor = conn.cursor()

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS trabajadores (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT NOT NULL,
        usuario TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        rol TEXT DEFAULT 'trabajador',
        activo INTEGER DEFAULT 1,
        fecha_creacion TEXT DEFAULT (datetime('now'))
    )
    """)

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS clientes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        placa TEXT UNIQUE NOT NULL,
        nombre TEXT,
        celular TEXT,
        precio_dia REAL,
        fecha_actualizacion TEXT
    )
    """)

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS entradas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cliente_id INTEGER,
        fecha_entrada TEXT,
        hora_entrada TEXT,
        fecha_hasta TEXT,
        hora_salida_esperada TEXT,
        hora_salida_real TEXT,
        dias INTEGER,
        precio_dia REAL,
        monto REAL,
        adelanto REAL DEFAULT 0,
        penalidad REAL DEFAULT 0,
        descuento REAL DEFAULT 0,
        metodo_pago TEXT DEFAULT 'efectivo',
        dejo_llave INTEGER DEFAULT 0,
        pagado INTEGER DEFAULT 0,
        pago_completo_adelantado INTEGER DEFAULT 0,
        salio INTEGER DEFAULT 0,
        observaciones TEXT,
        trabajador_id INTEGER,
        trabajador_salida_id INTEGER,
        fecha_registro TEXT,
        fecha_salida TEXT,
        FOREIGN KEY (cliente_id) REFERENCES clientes(id),
        FOREIGN KEY (trabajador_id) REFERENCES trabajadores(id),
        FOREIGN KEY (trabajador_salida_id) REFERENCES trabajadores(id)
    )
    """)
    
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS turnos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        trabajador_id INTEGER,
        fecha_inicio TEXT,
        fecha_fin TEXT,
        estado TEXT DEFAULT 'abierto',
        total_efectivo REAL DEFAULT 0,
        total_yape REAL DEFAULT 0,
        efectivo_declarado REAL,
        observaciones TEXT,
        FOREIGN KEY (trabajador_id) REFERENCES trabajadores(id)
    )
    """)
    
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS movimientos_caja (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        turno_id INTEGER,
        entrada_id INTEGER,
        trabajador_id INTEGER,
        tipo TEXT NOT NULL,
        monto REAL NOT NULL,
        metodo_pago TEXT DEFAULT 'efectivo',
        descripcion TEXT,
        fecha_movimiento TEXT DEFAULT (datetime('now')),
        FOREIGN KEY (turno_id) REFERENCES turnos(id),
        FOREIGN KEY (entrada_id) REFERENCES entradas(id),
        FOREIGN KEY (trabajador_id) REFERENCES trabajadores(id)
    )
    """)
    
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS configuracion (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        clave TEXT UNIQUE NOT NULL,
        valor TEXT,
        descripcion TEXT
    )
    """)
    
    configuraciones_default = [
        ('tolerancia_minutos', '60', 'Minutos de tolerancia antes de cobrar penalidad'),
        ('capacidad_maxima', '50', 'Capacidad máxima de la cochera'),
        ('precio_default', '10', 'Precio por día por defecto')
    ]
    
    for clave, valor, desc in configuraciones_default:
        cursor.execute("""
            INSERT OR IGNORE INTO configuracion (clave, valor, descripcion)
            VALUES (?, ?, ?)
        """, (clave, valor, desc))

    conn.commit()
    conn.close()


def crear_usuarios_default():
    conn = conectar_db()
    cursor = conn.cursor()
    
    usuarios = [
        ('Administrador', 'admin', 'admin123', 'admin'),
        ('Juan Pérez', 'juan', '1234', 'trabajador'),
        ('María García', 'maria', '1234', 'trabajador'),
        ('Carlos López', 'carlos', '1234', 'trabajador')
    ]
    
    for nombre, usuario, password, rol in usuarios:
        try:
            password_hash = generate_password_hash(password)
            cursor.execute("""
                INSERT INTO trabajadores (nombre, usuario, password, rol)
                VALUES (?, ?, ?, ?)
            """, (nombre, usuario, password_hash, rol))
        except sqlite3.IntegrityError:
            pass
    
    conn.commit()
    conn.close()


crear_tablas()
crear_usuarios_default()


# -------------------------------
# FUNCIONES AUXILIARES
# -------------------------------
def obtener_turno_activo(trabajador_id):
    conn = conectar_db()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT * FROM turnos 
        WHERE trabajador_id = ? AND estado = 'abierto'
        ORDER BY fecha_inicio DESC LIMIT 1
    """, (trabajador_id,))
    
    turno = cursor.fetchone()
    conn.close()
    
    return dict(turno) if turno else None


def crear_turno(trabajador_id):
    conn = conectar_db()
    cursor = conn.cursor()
    
    cursor.execute("""
        INSERT INTO turnos (trabajador_id, fecha_inicio, estado)
        VALUES (?, datetime('now'), 'abierto')
    """, (trabajador_id,))
    
    turno_id = cursor.lastrowid
    conn.commit()
    conn.close()
    
    return turno_id


def calcular_penalidad(fecha_entrada, hora_entrada, fecha_salida_esperada, hora_salida_esperada, precio_dia):
    try:
        conn = conectar_db()
        cursor = conn.cursor()
        cursor.execute("SELECT valor FROM configuracion WHERE clave = 'tolerancia_minutos'")
        result = cursor.fetchone()
        tolerancia = int(result['valor']) if result else 60
        conn.close()
        
        salida_esperada = datetime.strptime(f"{fecha_salida_esperada} {hora_salida_esperada}", "%Y-%m-%d %H:%M")
        ahora = datetime.now()
        salida_con_tolerancia = salida_esperada + timedelta(minutes=tolerancia)
        
        if ahora <= salida_con_tolerancia:
            return 0
        
        diferencia = ahora - salida_con_tolerancia
        horas_exceso = diferencia.total_seconds() / 3600
        precio_hora = float(precio_dia) / 24
        penalidad = round(horas_exceso * precio_hora, 2)
        
        return penalidad
        
    except Exception as e:
        print(f"Error calculando penalidad: {e}")
        return 0


def obtener_configuracion(clave, default=None):
    conn = conectar_db()
    cursor = conn.cursor()
    cursor.execute("SELECT valor FROM configuracion WHERE clave = ?", (clave,))
    result = cursor.fetchone()
    conn.close()
    return result['valor'] if result else default


# -------------------------------
# LOGIN
# -------------------------------
@app.route("/", methods=["GET", "POST"])
def login():
    error = None

    if request.method == "POST":
        usuario = request.form.get("usuario", "").strip()
        password = request.form.get("password", "")

        if not usuario or not password:
            error = "Usuario y contraseña son requeridos"
            return render_template("login.html", error=error)

        conn = None
        try:
            conn = conectar_db()
            cursor = conn.cursor()
            cursor.execute(
                "SELECT id, nombre, password, rol FROM trabajadores WHERE usuario=? AND activo=1",
                (usuario,)
            )
            trabajador = cursor.fetchone()

            if trabajador and check_password_hash(trabajador["password"], password):
                session.permanent = True
                session["trabajador_id"] = trabajador["id"]
                session["nombre"] = trabajador["nombre"]
                session["es_admin"] = trabajador["rol"] == "admin"
                session["rol"] = trabajador["rol"]
                
                turno = obtener_turno_activo(trabajador["id"])
                
                if turno:
                    session["turno_id"] = turno["id"]
                    session["inicio_turno"] = turno["fecha_inicio"]
                else:
                    turno_id = crear_turno(trabajador["id"])
                    session["turno_id"] = turno_id
                    session["inicio_turno"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                
                if trabajador["rol"] == "admin":
                    return redirect("/admin")
                else:
                    return redirect("/dashboard")
            else:
                error = "Usuario o contraseña incorrectos"

        except Exception as e:
            print(f"Error en login: {e}")
            error = "Error en el servidor"
        finally:
            if conn:
                conn.close()

    return render_template("login.html", error=error)


# -------------------------------
# DASHBOARD TRABAJADOR
# -------------------------------
@app.route("/dashboard")
@login_required
def dashboard():
    conn = None
    try:
        conn = conectar_db()
        cursor = conn.cursor()
        
        turno_id = session.get("turno_id")

        cursor.execute("""
            SELECT IFNULL(SUM(monto), 0)
            FROM movimientos_caja
            WHERE turno_id = ? AND metodo_pago = 'efectivo'
        """, (turno_id,))
        total_efectivo = cursor.fetchone()[0]
        
        cursor.execute("""
            SELECT IFNULL(SUM(monto), 0)
            FROM movimientos_caja
            WHERE turno_id = ? AND metodo_pago = 'yape'
        """, (turno_id,))
        total_yape = cursor.fetchone()[0]
        
        total_turno = total_efectivo + total_yape

        cursor.execute("""
            SELECT COUNT(*)
            FROM entradas
            WHERE trabajador_id = ?
            AND fecha_registro >= ?
        """, (session["trabajador_id"], session["inicio_turno"]))
        autos_ingresados = cursor.fetchone()[0]
        
        cursor.execute("""
            SELECT COUNT(*)
            FROM movimientos_caja
            WHERE turno_id = ? AND tipo = 'COBRO_SALIDA'
        """, (turno_id,))
        autos_salieron = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM entradas WHERE salio = 0")
        autos_en_cochera = cursor.fetchone()[0]

        return render_template(
            "dashboard.html",
            nombre=session["nombre"],
            total_turno=total_turno,
            total_efectivo=total_efectivo,
            total_yape=total_yape,
            autos_ingresados=autos_ingresados,
            autos_salieron=autos_salieron,
            autos_en_cochera=autos_en_cochera,
            inicio_turno=session["inicio_turno"],
            es_admin=session.get("es_admin", False)
        )

    except Exception as e:
        print(f"Error en dashboard: {e}")
        return "Error al cargar dashboard", 500
    finally:
        if conn:
            conn.close()


# -------------------------------
# DASHBOARD ADMIN
# -------------------------------
@app.route("/admin")
@admin_required
def admin_dashboard():
    conn = None
    try:
        conn = conectar_db()
        cursor = conn.cursor()
        
        hoy = datetime.now().strftime("%Y-%m-%d")
        mes_actual = datetime.now().strftime("%Y-%m")
        
        cursor.execute("""
            SELECT 
                IFNULL(SUM(CASE WHEN metodo_pago = 'efectivo' THEN monto ELSE 0 END), 0) as efectivo,
                IFNULL(SUM(CASE WHEN metodo_pago = 'yape' THEN monto ELSE 0 END), 0) as yape,
                IFNULL(SUM(monto), 0) as total
            FROM movimientos_caja
            WHERE date(fecha_movimiento) = ?
        """, (hoy,))
        ingresos_dia = cursor.fetchone()
        
        cursor.execute("""
            SELECT 
                IFNULL(SUM(CASE WHEN metodo_pago = 'efectivo' THEN monto ELSE 0 END), 0) as efectivo,
                IFNULL(SUM(CASE WHEN metodo_pago = 'yape' THEN monto ELSE 0 END), 0) as yape,
                IFNULL(SUM(monto), 0) as total
            FROM movimientos_caja
            WHERE strftime('%Y-%m', fecha_movimiento) = ?
        """, (mes_actual,))
        ingresos_mes = cursor.fetchone()
        
        cursor.execute("""
            SELECT 
                IFNULL(SUM(CASE WHEN metodo_pago = 'efectivo' THEN monto ELSE 0 END), 0) as efectivo,
                IFNULL(SUM(CASE WHEN metodo_pago = 'yape' THEN monto ELSE 0 END), 0) as yape,
                IFNULL(SUM(monto), 0) as total
            FROM movimientos_caja
        """)
        total_historico = cursor.fetchone()
        
        cursor.execute("SELECT COUNT(*) FROM entradas WHERE salio = 0")
        autos_en_cochera = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM clientes")
        total_clientes = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM trabajadores WHERE activo = 1")
        total_trabajadores = cursor.fetchone()[0]
        
        cursor.execute("""
            SELECT 
                date(fecha_movimiento) as fecha,
                SUM(monto) as total
            FROM movimientos_caja
            WHERE fecha_movimiento >= date('now', '-7 days')
            GROUP BY date(fecha_movimiento)
            ORDER BY fecha
        """)
        ingresos_semana = cursor.fetchall()

        return render_template(
            "admin_dashboard.html",
            nombre=session["nombre"],
            ingresos_dia=dict(ingresos_dia),
            ingresos_mes=dict(ingresos_mes),
            total_historico=dict(total_historico),
            autos_en_cochera=autos_en_cochera,
            total_clientes=total_clientes,
            total_trabajadores=total_trabajadores,
            ingresos_semana=[dict(i) for i in ingresos_semana]
        )

    except Exception as e:
        print(f"Error en admin dashboard: {e}")
        return "Error al cargar dashboard", 500
    finally:
        if conn:
            conn.close()


# -------------------------------
# GESTIÓN DE USUARIOS (ADMIN)
# -------------------------------
@app.route("/admin/usuarios")
@admin_required
def listar_usuarios():
    conn = conectar_db()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT id, nombre, usuario, rol, activo, fecha_creacion
        FROM trabajadores
        ORDER BY rol DESC, nombre
    """)
    
    usuarios = [dict(u) for u in cursor.fetchall()]
    conn.close()
    
    return jsonify({"ok": True, "usuarios": usuarios})


@app.route("/admin/usuarios/crear", methods=["POST"])
@admin_required
def crear_usuario():
    data = request.json
    
    if not data.get("nombre") or not data.get("usuario") or not data.get("password"):
        return jsonify({"ok": False, "error": "Todos los campos son requeridos"})
    
    conn = conectar_db()
    cursor = conn.cursor()
    
    try:
        password_hash = generate_password_hash(data["password"])
        cursor.execute("""
            INSERT INTO trabajadores (nombre, usuario, password, rol)
            VALUES (?, ?, ?, ?)
        """, (
            data["nombre"],
            data["usuario"].lower().strip(),
            password_hash,
            data.get("rol", "trabajador")
        ))
        
        conn.commit()
        return jsonify({"ok": True, "mensaje": "Usuario creado exitosamente"})
        
    except sqlite3.IntegrityError:
        return jsonify({"ok": False, "error": "El usuario ya existe"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})
    finally:
        conn.close()


@app.route("/admin/usuarios/editar", methods=["POST"])
@admin_required
def editar_usuario():
    data = request.json
    
    if not data.get("id"):
        return jsonify({"ok": False, "error": "ID requerido"})
    
    conn = conectar_db()
    cursor = conn.cursor()
    
    try:
        if data.get("password"):
            password_hash = generate_password_hash(data["password"])
            cursor.execute("""
                UPDATE trabajadores
                SET nombre = ?, usuario = ?, password = ?, rol = ?, activo = ?
                WHERE id = ?
            """, (
                data["nombre"],
                data["usuario"].lower().strip(),
                password_hash,
                data.get("rol", "trabajador"),
                1 if data.get("activo", True) else 0,
                data["id"]
            ))
        else:
            cursor.execute("""
                UPDATE trabajadores
                SET nombre = ?, usuario = ?, rol = ?, activo = ?
                WHERE id = ?
            """, (
                data["nombre"],
                data["usuario"].lower().strip(),
                data.get("rol", "trabajador"),
                1 if data.get("activo", True) else 0,
                data["id"]
            ))
        
        conn.commit()
        return jsonify({"ok": True, "mensaje": "Usuario actualizado"})
        
    except sqlite3.IntegrityError:
        return jsonify({"ok": False, "error": "El usuario ya existe"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})
    finally:
        conn.close()


@app.route("/admin/usuarios/eliminar/<int:id>", methods=["DELETE"])
@admin_required
def eliminar_usuario(id):
    if id == session["trabajador_id"]:
        return jsonify({"ok": False, "error": "No puedes eliminarte a ti mismo"})
    
    conn = conectar_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("UPDATE trabajadores SET activo = 0 WHERE id = ?", (id,))
        conn.commit()
        return jsonify({"ok": True, "mensaje": "Usuario desactivado"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})
    finally:
        conn.close()


# -------------------------------
# MOVIMIENTOS DEL TURNO
# -------------------------------
@app.route("/ingresos_turno")
@login_required
def ingresos_turno():
    conn = None
    try:
        conn = conectar_db()
        cursor = conn.cursor()
        
        turno_id = session.get("turno_id")

        cursor.execute("""
            SELECT 
                m.id,
                m.tipo,
                m.monto,
                m.metodo_pago,
                m.descripcion,
                m.fecha_movimiento,
                m.entrada_id,
                c.placa,
                c.nombre AS cliente
            FROM movimientos_caja m
            LEFT JOIN entradas e ON m.entrada_id = e.id
            LEFT JOIN clientes c ON e.cliente_id = c.id
            WHERE m.turno_id = ?
            ORDER BY m.fecha_movimiento DESC
        """, (turno_id,))

        rows = cursor.fetchall()

        ingresos = []
        total_efectivo = 0
        total_yape = 0

        for r in rows:
            monto = float(r["monto"] or 0)
            
            if r["metodo_pago"] == "efectivo":
                total_efectivo += monto
            else:
                total_yape += monto

            ingresos.append({
                "id": r["id"],
                "entrada_id": r["entrada_id"],
                "tipo": r["tipo"],
                "hora": r["fecha_movimiento"].split(" ")[1][:5] if r["fecha_movimiento"] else "",
                "fecha": r["fecha_movimiento"].split(" ")[0] if r["fecha_movimiento"] else "",
                "placa": r["placa"] or "-",
                "cliente": r["cliente"] or "-",
                "monto": monto,
                "metodo_pago": r["metodo_pago"],
                "descripcion": r["descripcion"]
            })

        return jsonify({
            "ingresos": ingresos,
            "total_efectivo": total_efectivo,
            "total_yape": total_yape,
            "total": total_efectivo + total_yape
        })

    except Exception as e:
        print(f"Error en ingresos_turno: {e}")
        return jsonify({"ingresos": [], "total": 0, "error": str(e)})
    finally:
        if conn:
            conn.close()


# -------------------------------
# BUSCAR CLIENTE
# -------------------------------
@app.route("/buscar_cliente/<placa>")
@login_required
def buscar_cliente(placa):
    conn = None
    try:
        conn = conectar_db()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT nombre, celular, precio_dia
            FROM clientes
            WHERE placa = ?
        """, (placa.upper().strip(),))

        cliente = cursor.fetchone()

        if cliente:
            return jsonify({
                "existe": True,
                "nombre": cliente["nombre"],
                "celular": cliente["celular"],
                "precio_dia": cliente["precio_dia"]
            })
        else:
            return jsonify({"existe": False})

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()


# -------------------------------
# GUARDAR ENTRADA
# -------------------------------
@app.route("/guardar_entrada", methods=["POST"])
@login_required
def guardar_entrada():
    data = request.json

    if not data.get("placa"):
        return jsonify({"ok": False, "error": "Placa es requerida"})

    if not data.get("cliente"):
        return jsonify({"ok": False, "error": "Nombre del cliente es requerido"})

    try:
        precio = float(data.get("precio", 0))
        if precio <= 0:
            return jsonify({"ok": False, "error": "El precio por día es obligatorio y debe ser mayor a 0"})
    except (ValueError, TypeError):
        return jsonify({"ok": False, "error": "Precio inválido"})

    try:
        dias = int(data.get("dias", 1))
        if dias < 1:
            return jsonify({"ok": False, "error": "Días debe ser al menos 1"})
    except (ValueError, TypeError):
        return jsonify({"ok": False, "error": "Días inválido"})

    conn = None
    try:
        conn = conectar_db()
        cursor = conn.cursor()

        placa = data["placa"].upper().strip()
        cursor.execute("SELECT id FROM clientes WHERE placa = ?", (placa,))
        cliente_db = cursor.fetchone()

        if cliente_db:
            cliente_id = cliente_db["id"]
            cursor.execute("""
                UPDATE clientes
                SET nombre=?, celular=?, precio_dia=?, fecha_actualizacion=datetime('now')
                WHERE id=?
            """, (data["cliente"], data.get("celular", ""), precio, cliente_id))
        else:
            cursor.execute("""
                INSERT INTO clientes (placa, nombre, celular, precio_dia, fecha_actualizacion)
                VALUES (?, ?, ?, ?, datetime('now'))
            """, (placa, data["cliente"], data.get("celular", ""), precio))
            cliente_id = cursor.lastrowid

        monto = precio * dias
        adelanto = float(data.get("adelanto", 0))
        metodo_pago = data.get("metodo_pago", "efectivo")
        
        pago_completo = 1 if data.get("pagado") and adelanto >= monto else 0
        
        if data.get("pagado") and adelanto == 0:
            adelanto = monto
            pago_completo = 1

        cursor.execute("""
            INSERT INTO entradas (
                cliente_id, fecha_entrada, hora_entrada,
                fecha_hasta, hora_salida_esperada, dias, precio_dia, monto, 
                adelanto, metodo_pago, dejo_llave, pagado, pago_completo_adelantado,
                salio, observaciones, trabajador_id, fecha_registro
            )
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,datetime('now'))
        """, (
            cliente_id,
            data.get("fecha_entrada"),
            data.get("hora_entrada"),
            data.get("fecha_hasta"),
            data.get("hora_salida"),
            dias,
            precio,
            monto,
            adelanto,
            metodo_pago,
            1 if data.get("dejo_llave") else 0,
            1 if pago_completo else 0,
            pago_completo,
            0,
            data.get("observaciones", ""),
            session["trabajador_id"]
        ))
        
        entrada_id = cursor.lastrowid

        if adelanto > 0:
            tipo_mov = "PAGO_COMPLETO" if pago_completo else "ADELANTO"
            cursor.execute("""
                INSERT INTO movimientos_caja (
                    turno_id, entrada_id, trabajador_id, tipo, monto, metodo_pago, descripcion
                )
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                session["turno_id"],
                entrada_id,
                session["trabajador_id"],
                tipo_mov,
                adelanto,
                metodo_pago,
                f"{tipo_mov} - {placa} - {data['cliente']} - {dias} día(s)"
            ))

        conn.commit()
        
        return jsonify({
            "ok": True, 
            "mensaje": "Entrada guardada exitosamente", 
            "id": entrada_id,
            "pago_completo": pago_completo == 1
        })

    except Exception as e:
        print(f"Error al guardar entrada: {e}")
        if conn:
            conn.rollback()
        return jsonify({"ok": False, "error": str(e)})
    finally:
        if conn:
            conn.close()


# -------------------------------
# OBTENER INGRESO POR ID
# -------------------------------
@app.route("/ingreso/<int:id>")
@login_required
def obtener_ingreso(id):
    conn = None
    try:
        conn = conectar_db()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT 
                e.*,
                c.placa,
                c.nombre as cliente,
                c.celular,
                t1.nombre as trabajador_entrada,
                t2.nombre as trabajador_salida
            FROM entradas e
            JOIN clientes c ON e.cliente_id = c.id
            LEFT JOIN trabajadores t1 ON e.trabajador_id = t1.id
            LEFT JOIN trabajadores t2 ON e.trabajador_salida_id = t2.id
            WHERE e.id = ?
        """, (id,))

        ingreso = cursor.fetchone()

        if ingreso:
            return jsonify(dict(ingreso))
        else:
            return jsonify({"error": "Ingreso no encontrado"}), 404

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()


# -------------------------------
# ACTUALIZAR INGRESO
# -------------------------------
@app.route("/actualizar_ingreso", methods=["POST"])
@login_required
def actualizar_ingreso():
    data = request.json

    if not data.get("id"):
        return jsonify({"ok": False, "error": "ID requerido"})

    conn = None
    try:
        conn = conectar_db()
        cursor = conn.cursor()

        cursor.execute("""
            UPDATE entradas
            SET fecha_entrada = ?, hora_entrada = ?, fecha_hasta = ?, 
                hora_salida_esperada = ?, precio_dia = ?, dias = ?, monto = ?,
                dejo_llave = ?, observaciones = ?
            WHERE id = ?
        """, (
            data.get("fecha_entrada"),
            data.get("hora_entrada"),
            data.get("fecha_hasta"),
            data.get("hora_salida"),
            float(data.get("precio", 0)),
            int(data.get("dias", 1)),
            float(data.get("monto", 0)),
            1 if data.get("dejo_llave") else 0,
            data.get("observaciones", ""),
            data["id"]
        ))

        if data.get("placa"):
            cursor.execute("""
                UPDATE clientes
                SET nombre = ?, celular = ?, precio_dia = ?
                WHERE placa = ?
            """, (
                data.get("cliente", ""),
                data.get("celular", ""),
                float(data.get("precio", 0)),
                data["placa"].upper().strip()
            ))

        conn.commit()
        return jsonify({"ok": True, "mensaje": "Ingreso actualizado"})

    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"ok": False, "error": str(e)})
    finally:
        if conn:
            conn.close()


# -------------------------------
# AUTOS EN COCHERA
# -------------------------------
@app.route("/autos_en_cochera")
@login_required
def autos_en_cochera():
    conn = None
    try:
        conn = conectar_db()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT 
                e.id,
                c.placa,
                c.nombre as cliente,
                c.celular,
                e.fecha_entrada,
                e.hora_entrada,
                e.fecha_hasta,
                e.hora_salida_esperada,
                e.dias as dias_pactados,
                e.precio_dia,
                e.monto,
                e.adelanto,
                e.metodo_pago,
                e.dejo_llave,
                e.pagado,
                e.pago_completo_adelantado,
                e.observaciones,
                t.nombre as trabajador_entrada,
                CAST((julianday('now') - julianday(e.fecha_entrada)) AS INTEGER) + 1 as dias_reales
            FROM entradas e
            JOIN clientes c ON e.cliente_id = c.id
            LEFT JOIN trabajadores t ON e.trabajador_id = t.id
            WHERE e.salio = 0
            ORDER BY e.fecha_entrada DESC
        """)

        autos = cursor.fetchall()

        autos_list = []
        for auto in autos:
            dias_reales = auto["dias_reales"]
            dias_pactados = auto["dias_pactados"]
            excede_tiempo = dias_reales > dias_pactados
            
            penalidad = 0
            if auto["fecha_hasta"] and auto["hora_salida_esperada"]:
                penalidad = calcular_penalidad(
                    auto["fecha_entrada"],
                    auto["hora_entrada"],
                    auto["fecha_hasta"],
                    auto["hora_salida_esperada"],
                    auto["precio_dia"]
                )
            
            monto_total = dias_reales * float(auto["precio_dia"]) + penalidad
            adelanto = float(auto["adelanto"] or 0)
            pendiente = max(0, monto_total - adelanto)

            autos_list.append({
                "id": auto["id"],
                "placa": auto["placa"],
                "cliente": auto["cliente"],
                "celular": auto["celular"],
                "fecha_entrada": auto["fecha_entrada"],
                "hora_entrada": auto["hora_entrada"],
                "fecha_hasta": auto["fecha_hasta"],
                "hora_salida_esperada": auto["hora_salida_esperada"],
                "dias_pactados": dias_pactados,
                "dias_reales": dias_reales,
                "precio_dia": auto["precio_dia"],
                "monto": auto["monto"],
                "adelanto": adelanto,
                "penalidad": penalidad,
                "pendiente": pendiente,
                "metodo_pago": auto["metodo_pago"],
                "dejo_llave": auto["dejo_llave"],
                "pagado": auto["pagado"],
                "pago_completo_adelantado": auto["pago_completo_adelantado"],
                "observaciones": auto["observaciones"],
                "trabajador_entrada": auto["trabajador_entrada"],
                "excede_tiempo": excede_tiempo
            })

        return jsonify({
            "ok": True,
            "autos": autos_list,
            "total": len(autos_list)
        })

    except Exception as e:
        print(f"Error en autos_en_cochera: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500
    finally:
        if conn:
            conn.close()


# -------------------------------
# CALCULAR COBRO
# -------------------------------
@app.route("/calcular_cobro/<int:id>")
@login_required
def calcular_cobro(id):
    conn = None
    try:
        conn = conectar_db()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT 
                e.*,
                c.placa,
                c.nombre as cliente,
                c.celular,
                t.nombre as trabajador_entrada,
                CAST((julianday('now') - julianday(e.fecha_entrada)) AS INTEGER) + 1 as dias_reales
            FROM entradas e
            JOIN clientes c ON e.cliente_id = c.id
            LEFT JOIN trabajadores t ON e.trabajador_id = t.id
            WHERE e.id = ?
        """, (id,))

        auto = cursor.fetchone()

        if not auto:
            return jsonify({"ok": False, "error": "Entrada no encontrada"})

        if auto["salio"]:
            return jsonify({"ok": False, "error": "Este auto ya salió"})

        dias_reales = auto["dias_reales"]
        precio_dia = float(auto["precio_dia"])
        adelanto = float(auto["adelanto"] or 0)
        
        penalidad = 0
        if auto["fecha_hasta"] and auto["hora_salida_esperada"]:
            penalidad = calcular_penalidad(
                auto["fecha_entrada"],
                auto["hora_entrada"],
                auto["fecha_hasta"],
                auto["hora_salida_esperada"],
                precio_dia
            )
        
        monto_dias = dias_reales * precio_dia
        monto_total = monto_dias + penalidad
        a_cobrar = max(0, monto_total - adelanto)
        
        ya_pago_completo = auto["pago_completo_adelantado"] == 1

        return jsonify({
            "ok": True,
            "id": auto["id"],
            "placa": auto["placa"],
            "cliente": auto["cliente"],
            "celular": auto["celular"],
            "trabajador_entrada": auto["trabajador_entrada"],
            "fecha_entrada": auto["fecha_entrada"],
            "hora_entrada": auto["hora_entrada"],
            "fecha_hasta": auto["fecha_hasta"],
            "hora_salida_esperada": auto["hora_salida_esperada"],
            "dias_pactados": auto["dias"],
            "dias_reales": dias_reales,
            "precio_dia": precio_dia,
            "monto_dias": monto_dias,
            "penalidad": penalidad,
            "monto_total": monto_total,
            "adelanto": adelanto,
            "a_cobrar": a_cobrar,
            "dejo_llave": auto["dejo_llave"],
            "ya_pago_completo": ya_pago_completo,
            "observaciones": auto["observaciones"]
        })

    except Exception as e:
        print(f"Error al calcular cobro: {e}")
        return jsonify({"ok": False, "error": str(e)})
    finally:
        if conn:
            conn.close()


# -------------------------------
# REGISTRAR SALIDA
# -------------------------------
@app.route("/registrar_salida", methods=["POST"])
@login_required
def registrar_salida():
    data = request.json

    if not data.get("id"):
        return jsonify({"ok": False, "error": "ID de entrada requerido"})

    conn = None
    try:
        conn = conectar_db()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT 
                e.*,
                c.placa,
                c.nombre as cliente_nombre,
                CAST((julianday('now') - julianday(e.fecha_entrada)) AS INTEGER) + 1 as dias_reales
            FROM entradas e
            JOIN clientes c ON e.cliente_id = c.id
            WHERE e.id = ?
        """, (data["id"],))

        entrada = cursor.fetchone()

        if not entrada:
            return jsonify({"ok": False, "error": "Entrada no encontrada"})

        if entrada["salio"] == 1:
            return jsonify({"ok": False, "error": "Este auto ya salió"})

        dias_reales = data.get("dias_reales", entrada["dias_reales"])
        precio_dia = float(entrada["precio_dia"])
        adelanto = float(entrada["adelanto"] or 0)
        penalidad = float(data.get("penalidad", 0))
        descuento = float(data.get("descuento", 0))
        metodo_pago = data.get("metodo_pago", "efectivo")
        
        monto_dias = dias_reales * precio_dia
        monto_total = monto_dias + penalidad - descuento
        monto_a_cobrar = max(0, monto_total - adelanto)

        if entrada["pago_completo_adelantado"] == 1:
            monto_a_cobrar = max(0, penalidad - descuento)

        cursor.execute("""
            UPDATE entradas
            SET salio = 1,
                fecha_salida = date('now'),
                hora_salida_real = time('now'),
                dias = ?,
                monto = ?,
                penalidad = ?,
                descuento = ?,
                pagado = 1,
                observaciones = ?,
                trabajador_salida_id = ?
            WHERE id = ?
        """, (
            dias_reales,
            monto_total,
            penalidad,
            descuento,
            data.get("observaciones", entrada["observaciones"]),
            session["trabajador_id"],
            data["id"]
        ))

        if monto_a_cobrar > 0:
            descripcion = f"Cobro salida - {entrada['placa']} - {entrada['cliente_nombre']} - {dias_reales} día(s)"
            if penalidad > 0:
                descripcion += f" (+ penalidad S/ {penalidad:.2f})"
            if descuento > 0:
                descripcion += f" (- descuento S/ {descuento:.2f})"
                
            cursor.execute("""
                INSERT INTO movimientos_caja (
                    turno_id, entrada_id, trabajador_id, tipo, monto, metodo_pago, descripcion
                )
                VALUES (?, ?, ?, 'COBRO_SALIDA', ?, ?, ?)
            """, (
                session["turno_id"],
                data["id"],
                session["trabajador_id"],
                monto_a_cobrar,
                metodo_pago,
                descripcion
            ))

        conn.commit()

        return jsonify({
            "ok": True,
            "mensaje": "Salida registrada exitosamente",
            "monto_total": monto_total,
            "adelanto": adelanto,
            "penalidad": penalidad,
            "descuento": descuento,
            "monto_cobrado": monto_a_cobrar,
            "dias_reales": dias_reales,
            "ya_pago_completo": entrada["pago_completo_adelantado"] == 1
        })

    except Exception as e:
        print(f"Error al registrar salida: {e}")
        if conn:
            conn.rollback()
        return jsonify({"ok": False, "error": str(e)})
    finally:
        if conn:
            conn.close()


# -------------------------------
# AUTORIZAR SALIDA (sin cobro)
# -------------------------------
@app.route("/autorizar_salida", methods=["POST"])
@login_required
def autorizar_salida():
    data = request.json

    if not data.get("id"):
        return jsonify({"ok": False, "error": "ID de entrada requerido"})

    conn = None
    try:
        conn = conectar_db()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT e.*, c.placa, c.nombre as cliente_nombre
            FROM entradas e
            JOIN clientes c ON e.cliente_id = c.id
            WHERE e.id = ? AND e.pago_completo_adelantado = 1 AND e.salio = 0
        """, (data["id"],))

        entrada = cursor.fetchone()

        if not entrada:
            return jsonify({"ok": False, "error": "Entrada no encontrada o no válida para autorización"})

        penalidad = float(data.get("penalidad", 0))
        descuento = float(data.get("descuento", 0))
        monto_extra = max(0, penalidad - descuento)
        metodo_pago = data.get("metodo_pago", "efectivo")

        cursor.execute("""
            UPDATE entradas
            SET salio = 1,
                fecha_salida = date('now'),
                hora_salida_real = time('now'),
                penalidad = ?,
                descuento = ?,
                observaciones = ?,
                trabajador_salida_id = ?
            WHERE id = ?
        """, (
            penalidad,
            descuento,
            data.get("observaciones", entrada["observaciones"]),
            session["trabajador_id"],
            data["id"]
        ))

        if monto_extra > 0:
            cursor.execute("""
                INSERT INTO movimientos_caja (
                    turno_id, entrada_id, trabajador_id, tipo, monto, metodo_pago, descripcion
                )
                VALUES (?, ?, ?, 'PENALIDAD', ?, ?, ?)
            """, (
                session["turno_id"],
                data["id"],
                session["trabajador_id"],
                monto_extra,
                metodo_pago,
                f"Penalidad - {entrada['placa']} - {entrada['cliente_nombre']}"
            ))

        conn.commit()

        return jsonify({
            "ok": True,
            "mensaje": "Salida autorizada exitosamente",
            "penalidad_cobrada": monto_extra
        })

    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"ok": False, "error": str(e)})
    finally:
        if conn:
            conn.close()


# -------------------------------
# VERIFICAR CAPACIDAD
# -------------------------------
@app.route("/verificar_capacidad")
@login_required
def verificar_capacidad():
    capacidad = int(obtener_configuracion('capacidad_maxima', 50))

    conn = None
    try:
        conn = conectar_db()
        cursor = conn.cursor()

        cursor.execute("SELECT COUNT(*) FROM entradas WHERE salio = 0")
        ocupados = cursor.fetchone()[0]

        disponibles = capacidad - ocupados
        porcentaje = (ocupados / capacidad) * 100

        return jsonify({
            "ok": True,
            "ocupados": ocupados,
            "disponibles": disponibles,
            "capacidad": capacidad,
            "porcentaje": round(porcentaje, 1)
        })

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})
    finally:
        if conn:
            conn.close()


# -------------------------------
# HISTORIAL COMPLETO
# -------------------------------
@app.route("/historial_vehiculos")
@login_required
def historial_vehiculos():
    conn = None
    try:
        conn = conectar_db()
        cursor = conn.cursor()
        
        filtro_placa = request.args.get('placa', '').upper()
        filtro_fecha_desde = request.args.get('fecha_desde', '')
        filtro_fecha_hasta = request.args.get('fecha_hasta', '')
        filtro_estado = request.args.get('estado', '')
        pagina = int(request.args.get('pagina', 1))
        por_pagina = int(request.args.get('por_pagina', 50))
        
        query = """
            SELECT 
                e.id,
                c.placa,
                c.nombre as cliente,
                c.celular,
                e.fecha_entrada,
                e.hora_entrada,
                e.fecha_salida,
                e.hora_salida_real,
                e.dias,
                e.precio_dia,
                e.monto,
                e.adelanto,
                e.penalidad,
                e.descuento,
                e.metodo_pago,
                e.pagado,
                e.pago_completo_adelantado,
                e.salio,
                e.observaciones,
                t1.nombre as trabajador_entrada,
                t2.nombre as trabajador_salida,
                CASE 
                    WHEN e.salio = 0 THEN CAST((julianday('now') - julianday(e.fecha_entrada)) AS INTEGER) + 1
                    ELSE e.dias
                END as dias_reales
            FROM entradas e
            JOIN clientes c ON e.cliente_id = c.id
            LEFT JOIN trabajadores t1 ON e.trabajador_id = t1.id
            LEFT JOIN trabajadores t2 ON e.trabajador_salida_id = t2.id
            WHERE 1=1
        """
        params = []
        
        if filtro_placa:
            query += " AND c.placa LIKE ?"
            params.append(f"%{filtro_placa}%")
        
        if filtro_fecha_desde:
            query += " AND e.fecha_entrada >= ?"
            params.append(filtro_fecha_desde)
        
        if filtro_fecha_hasta:
            query += " AND e.fecha_entrada <= ?"
            params.append(filtro_fecha_hasta)
        
        if filtro_estado == 'en_cochera':
            query += " AND e.salio = 0"
        elif filtro_estado == 'salieron':
            query += " AND e.salio = 1"
        
        count_query = f"SELECT COUNT(*) FROM ({query})"
        cursor.execute(count_query, params)
        total = cursor.fetchone()[0]
        
        query += " ORDER BY e.fecha_entrada DESC, e.hora_entrada DESC"
        query += f" LIMIT {por_pagina} OFFSET {(pagina - 1) * por_pagina}"
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        
        historial = []
        for r in rows:
            historial.append({
                "id": r["id"],
                "placa": r["placa"],
                "cliente": r["cliente"],
                "celular": r["celular"],
                "fecha_entrada": r["fecha_entrada"],
                "hora_entrada": r["hora_entrada"],
                "fecha_salida": r["fecha_salida"],
                "hora_salida": r["hora_salida_real"],
                "dias": r["dias"],
                "dias_reales": r["dias_reales"],
                "precio_dia": r["precio_dia"],
                "monto": r["monto"],
                "adelanto": r["adelanto"],
                "penalidad": r["penalidad"],
                "descuento": r["descuento"],
                "metodo_pago": r["metodo_pago"],
                "pagado": r["pagado"],
                "pago_completo": r["pago_completo_adelantado"],
                "salio": r["salio"],
                "observaciones": r["observaciones"],
                "trabajador_entrada": r["trabajador_entrada"],
                "trabajador_salida": r["trabajador_salida"]
            })
        
        return jsonify({
            "ok": True,
            "historial": historial,
            "total": total,
            "pagina": pagina,
            "por_pagina": por_pagina,
            "total_paginas": (total + por_pagina - 1) // por_pagina
        })
        
    except Exception as e:
        print(f"Error en historial: {e}")
        return jsonify({"ok": False, "error": str(e)})
    finally:
        if conn:
            conn.close()


# -------------------------------
# EXPORTAR A EXCEL
# -------------------------------
@app.route("/exportar_historial")
@login_required
def exportar_historial():
    try:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return jsonify({"ok": False, "error": "Módulo openpyxl no instalado"})
        
        conn = conectar_db()
        cursor = conn.cursor()
        
        filtro_placa = request.args.get('placa', '').upper()
        filtro_fecha_desde = request.args.get('fecha_desde', '')
        filtro_fecha_hasta = request.args.get('fecha_hasta', '')
        filtro_estado = request.args.get('estado', '')
        
        query = """
            SELECT 
                c.placa,
                c.nombre as cliente,
                c.celular,
                e.fecha_entrada,
                e.hora_entrada,
                e.fecha_salida,
                e.hora_salida_real as hora_salida,
                e.dias,
                e.precio_dia,
                e.monto,
                e.adelanto,
                e.penalidad,
                e.descuento,
                e.metodo_pago,
                CASE WHEN e.salio = 1 THEN 'Sí' ELSE 'No' END as salio,
                CASE WHEN e.pagado = 1 THEN 'Sí' ELSE 'No' END as pagado,
                t1.nombre as trabajador_entrada,
                t2.nombre as trabajador_salida,
                e.observaciones
            FROM entradas e
            JOIN clientes c ON e.cliente_id = c.id
            LEFT JOIN trabajadores t1 ON e.trabajador_id = t1.id
            LEFT JOIN trabajadores t2 ON e.trabajador_salida_id = t2.id
            WHERE 1=1
        """
        params = []
        
        if filtro_placa:
            query += " AND c.placa LIKE ?"
            params.append(f"%{filtro_placa}%")
        if filtro_fecha_desde:
            query += " AND e.fecha_entrada >= ?"
            params.append(filtro_fecha_desde)
        if filtro_fecha_hasta:
            query += " AND e.fecha_entrada <= ?"
            params.append(filtro_fecha_hasta)
        if filtro_estado == 'en_cochera':
            query += " AND e.salio = 0"
        elif filtro_estado == 'salieron':
            query += " AND e.salio = 1"
        
        query += " ORDER BY e.fecha_entrada DESC"
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        conn.close()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Historial"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = ['Placa', 'Cliente', 'Celular', 'F. Entrada', 'H. Entrada', 
                   'F. Salida', 'H. Salida', 'Días', 'Precio/Día', 'Monto Total',
                   'Adelanto', 'Penalidad', 'Descuento', 'Método Pago', 'Salió', 
                   'Pagado', 'Trabajador Entrada', 'Trabajador Salida', 'Observaciones']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(dict(row).values(), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 30)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=historial_cochera_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        print(f"Error exportando: {e}")
        return jsonify({"ok": False, "error": str(e)})


# -------------------------------
# HISTORIAL DEL CLIENTE
# -------------------------------
@app.route("/historial_cliente/<placa>")
@login_required
def historial_cliente(placa):
    conn = None
    try:
        conn = conectar_db()
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM clientes WHERE placa = ?", (placa.upper(),))
        cliente = cursor.fetchone()

        if not cliente:
            return jsonify({"ok": False, "error": "Cliente no encontrado"})

        cursor.execute("""
            SELECT 
                e.*,
                t1.nombre as trabajador_entrada,
                t2.nombre as trabajador_salida
            FROM entradas e
            LEFT JOIN trabajadores t1 ON e.trabajador_id = t1.id
            LEFT JOIN trabajadores t2 ON e.trabajador_salida_id = t2.id
            WHERE e.cliente_id = ?
            ORDER BY e.fecha_registro DESC
        """, (cliente["id"],))

        visitas = cursor.fetchall()

        cursor.execute("""
            SELECT 
                COUNT(*) as total_visitas,
                SUM(monto) as total_gastado,
                SUM(CASE WHEN pagado = 0 THEN monto - IFNULL(adelanto, 0) ELSE 0 END) as deuda_actual,
                AVG(dias) as promedio_dias
            FROM entradas
            WHERE cliente_id = ?
        """, (cliente["id"],))

        estadisticas = cursor.fetchone()

        return jsonify({
            "ok": True,
            "cliente": dict(cliente),
            "estadisticas": dict(estadisticas),
            "visitas": [dict(v) for v in visitas]
        })

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})
    finally:
        if conn:
            conn.close()


# -------------------------------
# DETALLE DE MOVIMIENTO
# -------------------------------
@app.route("/detalle_movimiento/<int:id>")
@login_required
def detalle_movimiento(id):
    conn = None
    try:
        conn = conectar_db()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT 
                m.*,
                e.fecha_entrada,
                e.hora_entrada,
                e.fecha_salida,
                e.hora_salida_real,
                e.dias,
                e.precio_dia,
                e.monto as monto_total,
                e.adelanto,
                e.penalidad,
                e.descuento,
                e.observaciones,
                e.pago_completo_adelantado,
                c.placa,
                c.nombre as cliente,
                c.celular,
                t1.nombre as trabajador_entrada,
                t2.nombre as trabajador_salida,
                t3.nombre as trabajador_movimiento
            FROM movimientos_caja m
            LEFT JOIN entradas e ON m.entrada_id = e.id
            LEFT JOIN clientes c ON e.cliente_id = c.id
            LEFT JOIN trabajadores t1 ON e.trabajador_id = t1.id
            LEFT JOIN trabajadores t2 ON e.trabajador_salida_id = t2.id
            LEFT JOIN trabajadores t3 ON m.trabajador_id = t3.id
            WHERE m.id = ?
        """, (id,))

        mov = cursor.fetchone()

        if not mov:
            return jsonify({"ok": False, "error": "Movimiento no encontrado"})

        return jsonify({
            "ok": True,
            "movimiento": dict(mov)
        })

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})
    finally:
        if conn:
            conn.close()


# -------------------------------
# ALERTAS
# -------------------------------
@app.route("/obtener_alertas")
@login_required
def obtener_alertas():
    conn = None
    try:
        conn = conectar_db()
        cursor = conn.cursor()

        alertas = []

        cursor.execute("""
            SELECT 
                c.placa,
                c.nombre as cliente,
                e.dias as dias_pactados,
                CAST((julianday('now') - julianday(e.fecha_entrada)) AS INTEGER) + 1 as dias_reales
            FROM entradas e
            JOIN clientes c ON e.cliente_id = c.id
            WHERE e.salio = 0
            AND CAST((julianday('now') - julianday(e.fecha_entrada)) AS INTEGER) + 1 > e.dias
        """)

        for auto in cursor.fetchall():
            exceso = auto["dias_reales"] - auto["dias_pactados"]
            alertas.append({
                "tipo": "exceso_tiempo",
                "nivel": "warning",
                "titulo": f"⚠️ {auto['placa']} excede tiempo",
                "mensaje": f"{auto['cliente']} - Exceso: {exceso} día(s)",
                "placa": auto["placa"]
            })

        capacidad = int(obtener_configuracion('capacidad_maxima', 50))
        cursor.execute("SELECT COUNT(*) FROM entradas WHERE salio = 0")
        ocupados = cursor.fetchone()[0]
        porcentaje = (ocupados / capacidad) * 100

        if porcentaje >= 90:
            alertas.append({
                "tipo": "capacidad",
                "nivel": "danger",
                "titulo": "🚨 Cochera casi llena",
                "mensaje": f"Ocupación: {porcentaje:.0f}% ({ocupados}/{capacidad})",
                "placa": None
            })
        elif porcentaje >= 75:
            alertas.append({
                "tipo": "capacidad",
                "nivel": "warning",
                "titulo": "⚠️ Cochera llegando al límite",
                "mensaje": f"Ocupación: {porcentaje:.0f}% ({ocupados}/{capacidad})",
                "placa": None
            })

        return jsonify({
            "ok": True,
            "alertas": alertas,
            "total": len(alertas)
        })

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})
    finally:
        if conn:
            conn.close()


# -------------------------------
# GENERAR TICKET
# -------------------------------
@app.route("/generar_ticket/<int:id>")
@login_required
def generar_ticket(id):
    conn = None
    try:
        conn = conectar_db()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT 
                e.*,
                c.placa,
                c.nombre as cliente,
                c.celular,
                t.nombre as trabajador
            FROM entradas e
            JOIN clientes c ON e.cliente_id = c.id
            LEFT JOIN trabajadores t ON e.trabajador_id = t.id
            WHERE e.id = ?
        """, (id,))

        entrada = cursor.fetchone()

        if not entrada:
            return jsonify({"ok": False, "error": "Entrada no encontrada"})

        return jsonify({
            "ok": True,
            "ticket": {
                "id": entrada["id"],
                "placa": entrada["placa"],
                "cliente": entrada["cliente"],
                "fecha_entrada": entrada["fecha_entrada"],
                "hora_entrada": entrada["hora_entrada"],
                "fecha_hasta": entrada["fecha_hasta"],
                "hora_salida_esperada": entrada["hora_salida_esperada"],
                "dias": entrada["dias"],
                "precio_dia": entrada["precio_dia"],
                "monto": entrada["monto"],
                "adelanto": entrada["adelanto"],
                "trabajador": entrada["trabajador"],
                "fecha_emision": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
        })

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})
    finally:
        if conn:
            conn.close()


# -------------------------------
# REPORTE DEL TURNO
# -------------------------------
@app.route("/reporte_turno")
@login_required
def reporte_turno():
    conn = None
    try:
        conn = conectar_db()
        cursor = conn.cursor()
        
        turno_id = session.get("turno_id")

        cursor.execute("""
            SELECT 
                IFNULL(SUM(monto), 0) as total_cobrado,
                IFNULL(SUM(CASE WHEN metodo_pago = 'efectivo' THEN monto ELSE 0 END), 0) as total_efectivo,
                IFNULL(SUM(CASE WHEN metodo_pago = 'yape' THEN monto ELSE 0 END), 0) as total_yape,
                IFNULL(SUM(CASE WHEN tipo LIKE '%ADELANTO%' OR tipo = 'PAGO_COMPLETO' THEN monto ELSE 0 END), 0) as total_adelantos,
                IFNULL(SUM(CASE WHEN tipo = 'COBRO_SALIDA' THEN monto ELSE 0 END), 0) as total_cobros,
                IFNULL(SUM(CASE WHEN tipo = 'PENALIDAD' THEN monto ELSE 0 END), 0) as total_penalidades
            FROM movimientos_caja
            WHERE turno_id = ?
        """, (turno_id,))
        
        stats_mov = cursor.fetchone()

        cursor.execute("""
            SELECT COUNT(*) FROM entradas
            WHERE trabajador_id = ? AND fecha_registro >= ?
        """, (session["trabajador_id"], session["inicio_turno"]))
        autos_ingresados = cursor.fetchone()[0]
        
        cursor.execute("""
            SELECT COUNT(*) FROM movimientos_caja
            WHERE turno_id = ? AND tipo = 'COBRO_SALIDA'
        """, (turno_id,))
        autos_salieron = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM entradas WHERE salio = 0")
        autos_en_cochera = cursor.fetchone()[0]

        stats = {
            "total_cobrado": stats_mov["total_cobrado"],
            "total_efectivo": stats_mov["total_efectivo"],
            "total_yape": stats_mov["total_yape"],
            "total_adelantos": stats_mov["total_adelantos"],
            "total_cobros": stats_mov["total_cobros"],
            "total_penalidades": stats_mov["total_penalidades"],
            "autos_ingresados": autos_ingresados,
            "autos_salieron": autos_salieron,
            "autos_en_cochera": autos_en_cochera
        }

        cursor.execute("""
            SELECT 
                m.*,
                c.placa,
                c.nombre as cliente
            FROM movimientos_caja m
            LEFT JOIN entradas e ON m.entrada_id = e.id
            LEFT JOIN clientes c ON e.cliente_id = c.id
            WHERE m.turno_id = ?
            ORDER BY m.fecha_movimiento DESC
        """, (turno_id,))

        detalles = [dict(d) for d in cursor.fetchall()]

        return render_template(
            "reporte_turno.html",
            trabajador=session["nombre"],
            inicio_turno=session["inicio_turno"],
            fin_turno=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            stats=stats,
            detalles=detalles
        )

    except Exception as e:
        print(f"Error en reporte: {e}")
        return "Error al generar reporte", 500
    finally:
        if conn:
            conn.close()


# -------------------------------
# CERRAR TURNO
# -------------------------------
@app.route("/cerrar_turno", methods=["POST"])
@login_required
def cerrar_turno():
    data = request.json
    
    conn = None
    try:
        conn = conectar_db()
        cursor = conn.cursor()
        
        turno_id = session.get("turno_id")

        cursor.execute("""
            SELECT 
                IFNULL(SUM(CASE WHEN metodo_pago = 'efectivo' THEN monto ELSE 0 END), 0) as total_efectivo,
                IFNULL(SUM(CASE WHEN metodo_pago = 'yape' THEN monto ELSE 0 END), 0) as total_yape,
                IFNULL(SUM(monto), 0) as total
            FROM movimientos_caja
            WHERE turno_id = ?
        """, (turno_id,))

        totales = cursor.fetchone()
        
        cursor.execute("""
            SELECT COUNT(*) FROM entradas
            WHERE trabajador_id = ? AND fecha_registro >= ?
        """, (session["trabajador_id"], session["inicio_turno"]))
        autos_ingresados = cursor.fetchone()[0]
        
        cursor.execute("""
            SELECT COUNT(*) FROM movimientos_caja
            WHERE turno_id = ? AND tipo = 'COBRO_SALIDA'
        """, (turno_id,))
        autos_salieron = cursor.fetchone()[0]

        efectivo_declarado = float(data.get("efectivo_declarado", 0))
        total_efectivo = float(totales["total_efectivo"] or 0)
        diferencia = efectivo_declarado - total_efectivo

        if data.get("solo_calcular"):
            return jsonify({
                "ok": True,
                "autos_ingresados": autos_ingresados,
                "autos_salieron": autos_salieron,
                "total_efectivo": total_efectivo,
                "total_yape": float(totales["total_yape"] or 0),
                "total_cobrado": float(totales["total"] or 0),
                "efectivo_declarado": efectivo_declarado,
                "diferencia": diferencia
            })

        cursor.execute("""
            UPDATE turnos
            SET estado = 'cerrado',
                fecha_fin = datetime('now'),
                total_efectivo = ?,
                total_yape = ?,
                efectivo_declarado = ?,
                observaciones = ?
            WHERE id = ?
        """, (
            total_efectivo,
            float(totales["total_yape"] or 0),
            efectivo_declarado,
            data.get("observaciones", ""),
            turno_id
        ))
        
        conn.commit()

        return jsonify({
            "ok": True,
            "mensaje": "Turno cerrado exitosamente",
            "cerrado": True,
            "autos_ingresados": autos_ingresados,
            "autos_salieron": autos_salieron,
            "total_efectivo": total_efectivo,
            "total_yape": float(totales["total_yape"] or 0),
            "efectivo_declarado": efectivo_declarado,
            "diferencia": diferencia
        })

    except Exception as e:
        print(f"Error al cerrar turno: {e}")
        return jsonify({"ok": False, "error": str(e)})
    finally:
        if conn:
            conn.close()


# -------------------------------
# CONFIGURACIÓN
# -------------------------------
@app.route("/admin/configuracion")
@admin_required
def obtener_configuracion_all():
    conn = conectar_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM configuracion")
    config = [dict(c) for c in cursor.fetchall()]
    conn.close()
    return jsonify({"ok": True, "configuracion": config})


@app.route("/admin/configuracion", methods=["POST"])
@admin_required
def guardar_configuracion_admin():
    data = request.json
    
    conn = conectar_db()
    cursor = conn.cursor()
    
    try:
        for clave, valor in data.items():
            cursor.execute("""
                UPDATE configuracion SET valor = ? WHERE clave = ?
            """, (valor, clave))
        
        conn.commit()
        return jsonify({"ok": True, "mensaje": "Configuración guardada"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})
    finally:
        conn.close()


# -------------------------------
# LOGOUT
# -------------------------------
@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")


# -------------------------------
# EJECUTAR APP
# -------------------------------
if __name__ == "__main__":
    # En desarrollo
    app.run(debug=True, host='0.0.0.0', port=5000)
